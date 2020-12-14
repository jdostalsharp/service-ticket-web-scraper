#! python 3

"""
Web scraper to get the USB device information from ##### [removed] ##### Tasks and save in an excel file in the correct format.

Author: Jeremy Dostal-Sharp
Intern Cyber Security self designated tasks
"""

import sys
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import bs4

# Excel file name to be created
SPREADSHEET_NAME = "New USB devices.xlsx"

# Service-now service catalog website address
URL = "www.google.com" # Proper URL Removed for security reasons

class CreateExcelWorkbook:
    """
    Create an excel spreadsheet with file name SPREADSHEET_NAME and sheet name "New USB List".
    """

    def __init__(self):
        self.workbook = Workbook()
        self.dest_filename = SPREADSHEET_NAME
        self.worksheet = self.workbook.active
        self.worksheet.title = "New USB List"
        self.row = 2
        # Wipe any data that may be on the sheet already
        for row in self.worksheet['A1:J100']:
            for cell in row:
                cell.value = None
        self.worksheet.cell(row=1, column=1).value = "Tickets that seem correct"

    def data_sorting(self, device_list):
        """
        Sorts the devices between good and bad devices depending on their model code length
        :param device_list: List of devices which are a list (list of lists)
        :return: none
        """
        bad_device = []
        for device in device_list:
            if len(device) == 10:
                # Weed out the bad devices
                if self.data_check(device[2]):
                    self.add_usb_info(device[0], device[1], device[2], device[3], device[4], device[5], device[6], device[7], device[8], device[9])
                else:
                    bad_device.append(device)
            else:
                print("Something went wrong with the device list")

        # Add the bad devices at the bottom of the list so they can be checked manually
        self.row += 2
        self.worksheet.cell(row=self.row, column=1).value = "Tickets with incorrect values"
        self.row += 1
        for device in bad_device:
            self.add_usb_info(device[0], device[1], device[2], device[3], device[4], device[5], device[6], device[7],
                              device[8], device[9])

        self.workbook.save(self.dest_filename)

    def add_usb_info(self, ticket_no, vendor, model, serial, device_type, source, machine_name, email, username,
                     justification):
        """
        Add item to spreadsheet in the correct layout
        """
        self.worksheet.cell(row=self.row, column=1).value = ticket_no
        self.worksheet.cell(row=self.row, column=2).value = vendor
        self.worksheet.cell(row=self.row, column=3).value = model
        self.worksheet.cell(row=self.row, column=4).value = serial
        self.worksheet.cell(row=self.row, column=5).value = device_type
        self.worksheet.cell(row=self.row, column=6).value = source
        self.worksheet.cell(row=self.row, column=7).value = machine_name
        self.worksheet.cell(row=self.row, column=8).value = email
        self.worksheet.cell(row=self.row, column=9).value = username
        self.worksheet.cell(row=self.row, column=10).value = justification
        self.row += 1

    def data_check(self, model):
        """
        Checks that model has four characters
        :param model: string
        :return: True or False
        """
        if len(model) == 4:
            return True
        else:
            return False


class DataScarper:
    """
    Download the info from each ticket with
        short description = Add USB to allowed USB list
        active = True
    From ## [Removed] ## Service_Catalog Tasks
    """
    def __init__(self):
        self.service_data = ''
        self.device_list = []

    """
    Opens the web browser and runs the webscraping function    
    """
    def get_data(self):
        # Open the browser to the required page
        browser = webdriver.Edge()
        browser.get(URL)

        # This waits for input allows the user the chance to login securely, rather than save their
        # login details in the code or a plaintext file
        input('Please login to the website and press enter here when service now page is loaded')

        # Switch to the correct iframe on the website
        browser.switch_to.frame(browser.find_element_by_id('gsft_main'))

        more_tickets = 'y'
        while more_tickets == 'yes' or more_tickets == 'y':
            self.scrape_page(browser)
            more_tickets = 'no'
            print("Is there another page of tickets? If so change the page and type 'yes', otherwise type 'no")
            more_tickets = input()
            more_tickets.lower()

        return self.device_list

    """
    Scrapes the page running get_device_data as many times as needed
    """
    def scrape_page(self, browser):
        # Set the iteration counter and the total tickets on that page
        iteration = 0
        total_tickets = len(browser.find_elements_by_partial_link_text("TASK0"))
        # Max iterations has been set to 20 as that is how many
        while iteration < total_tickets:
            # Print statement for testing
            print("Starting data extraction")

            # Get a list of all the elements we want to interact with
            tasks = browser.find_elements_by_partial_link_text("TASK0")

            tasks[iteration].click()
            self.service_data = browser.page_source
            device_info = self.get_device_data(browser)
            for device in device_info:
                self.device_list.append(device)

            # Currently not completing tickets just pressing the back button
            back = browser.find_elements_by_xpath(
                '//*[@id="section-8687fbccc611229100727249a775cc31.header"]/nav/div/div[1]/button[1]')
            back[0].click()
            iteration += 1

    """
    Gets the device data for each device on a ticket and adds it to a list
    """
    def get_device_data(self, browser):
        soup = bs4.BeautifulSoup(self.service_data, 'html.parser')
        elems = soup.select('td div > div > div > div > table > tbody > tr > td')
        line_entries = []
        # print(str(elems))
        # Create a entry for the excel file with data [Vendor, Model, Serial id, Device Type]
        iterations = 0
        line_entry = ['', '', '', '', '', '', '', '', '', '']
        length = len(elems)

        # Get the page source info from the pop up
        info_button = browser.find_elements_by_css_selector('#viewr\.sc_task\.request_item\.request\.u_requested_by')
        info_button[0].click()
        # The wait has been added to allow time to get the info page up for ach item
        WebDriverWait(browser, 20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, '#sys_user\.do')))
        info_button_soup = bs4.BeautifulSoup(browser.page_source, 'html.parser')

        while iterations < length:
            if iterations % 5 == 0:
                # print("we are here")
                elems.pop(0)
                line_entry = ['', '', '', '', '', '', '', '', '', '']
                iterations += 1
            else:
                # Get the main info
                device_type = elems.pop(0).getText()
                # print(device_type)
                vendor = elems.pop(0).getText()
                # print(vendor)
                model = elems.pop(0).getText()
                # print(model)
                serial = elems.pop(0).getText()
                # print(serial)
                line_entry[4] = device_type
                line_entry[1] = vendor
                line_entry[2] = model
                line_entry[3] = serial
                # get to the next lot of values
                iterations += 4

                # Get the ticket number
                ticket_no = soup.select('#sys_readonly\.sc_task\.number')
                if len(ticket_no) > 0:
                    line_entry[0] = ticket_no[0].attrs['value']
                    # print(ticket_no[0].attrs['value'])

                # Get the source/Reference field and append it to the list
                source = soup.select('#sys_display\.sc_task\.request_item')
                if len(source) > 0:
                    line_entry[5] = source[0].attrs['value']
                    # print(source[0].attrs['value'])

                # Get the Machine name data and append it to the list
                mac_name = soup.select('td > div > div > table > tbody > tr > td > div > div > div > div input')
                if len(mac_name) > 1:
                    line_entry[6] = mac_name[1].attrs['value']
                    # print(mac_name[1].attrs['value'])

                # Get the email and from the pop up
                email = info_button_soup.select('#sys_readonly\.sys_user\.email')
                if len(email) > 0:
                    line_entry[7] = email[0].attrs['value']

                # Get the username from the pop up
                username_full = info_button_soup.select('#sys_readonly\.sys_user\.user_name')
                if len(username_full) > 0:
                    temp = username_full[0].attrs['value']
                    username = temp.split("@")[0]
                    line_entry[8] = username

                # Get the justification data and append it to the list
                just = soup.select('tbody > tr > td > div > div > div > div > textarea')
                if len(just) > 0:
                    line_entry[9] = just[1].getText()
                    # print(just[0].getText())

                line_entries.append(line_entry.copy())
                # print("About to save: ")
                # print(line_entry)

        # Printing out the string for testing purposes
        # print(line_entries)
        print("Data retrieved from ticket " + ticket_no[0].attrs['value'])
        return line_entries


def main(arglist):
    """
    Runs the Program
    :param arglist: none
    :return: Saves an excel document in the directory run with the name defined in SPREADSHEET_NAME
    """
    current_workbook = CreateExcelWorkbook()
    scraper = DataScarper()
    current_workbook.data_sorting(scraper.get_data())
    print("Please remove the data from the new excel spreadsheet and close the spreadsheet before running again")
    input('When you are done, press enter key to close the browser')


if __name__ == '__main__':
    main(sys.argv[1:])
